import os
import pandas as pd

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
PERIODS = list(range(1, 11))


def _cell_text(a):
    return (
        f"{a.subject.name}\n"
        f"GV: {a.teacher.name}\n"
        f"Phòng: {a.room.id}"
    )


def _format_sheet(ws):
    header_fill = PatternFill(
        fill_type="solid",
        start_color="4F81BD",
        end_color="4F81BD"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin = Side(style="thin", color="000000")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = border

            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    ws.column_dimensions["A"].width = 8

    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 28

    for row_idx in range(2, 12):
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "B2"


def _format_all_assignments(ws):
    header_fill = PatternFill(
        fill_type="solid",
        start_color="70AD47",
        end_color="70AD47"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin = Side(style="thin", color="000000")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = border

            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_schedule(assignments, output_path="output/timetable.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    raw_rows = []

    for a in assignments:
        raw_rows.append({
            "Lớp": a.class_id,
            "Mã môn": a.subject.id,
            "Môn": a.subject.name,
            "Mã GV": a.teacher.id,
            "Giáo viên": a.teacher.name,
            "Phòng": a.room.id,
            "Thứ": a.session.day,
            "Tiết": a.session.period,
        })

    raw_df = pd.DataFrame(raw_rows)

    if not raw_df.empty:
        raw_df = raw_df.sort_values(
            by=["Lớp", "Thứ", "Tiết"]
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw_df.to_excel(
            writer,
            sheet_name="All_Assignments",
            index=False
        )

        _format_all_assignments(writer.sheets["All_Assignments"])

        if raw_df.empty:
            print("Không có assignment để xuất.")
            return

        for class_id in sorted(raw_df["Lớp"].unique()):
            table = pd.DataFrame(
                index=PERIODS,
                columns=DAYS
            )

            table.index.name = "Tiết"

            class_assignments = [
                a for a in assignments
                if a.class_id == class_id
            ]

            for a in class_assignments:
                table.loc[a.session.period, a.session.day] = _cell_text(a)

            table.to_excel(
                writer,
                sheet_name=str(class_id)
            )

            ws = writer.sheets[str(class_id)]
            _format_sheet(ws)

    print(f"Đã xuất thời khóa biểu: {output_path}")